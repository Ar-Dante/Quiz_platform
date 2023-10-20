import io

from fastapi import HTTPException
from openpyxl import load_workbook

from app.conf.messages import ERROR_NOT_EXCEL_FORMAT
from app.repository.questions import QuestionsRepository
from app.repository.quizzes import QuizzesRepository
from app.schemas.questions_schemas import QuestionCreateModel
from app.schemas.quizzes_schemas import QuizCreateModel, QuizUpdateModel
from app.services.questions import QuestionService
from app.services.quizzes import QuizService

quizzes_service = QuizService(QuizzesRepository)
question_service = QuestionService(QuestionsRepository)


async def import_quiz_from_excel(file, company: dict, member: dict, current_user: int):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=ERROR_NOT_EXCEL_FORMAT)

    file_data = await file.read()
    wb = load_workbook((io.BytesIO(file_data)), data_only=True)
    sheet = wb.active

    quiz_name = sheet['A1'].value
    quiz_title = sheet['A2'].value
    quiz_description = sheet['A3'].value
    quiz_frequency = sheet['A4'].value

    quiz_data = QuizCreateModel(
        quiz_name=quiz_name,
        quiz_description=quiz_description,
        quiz_title=quiz_title,
        quiz_frequency=quiz_frequency
    )

    quiz_id = await quizzes_service.create_quiz(company, quiz_data, member, current_user)
    await create_questions_from_excel(quiz_id, sheet, member, company, current_user)


async def update_quiz_from_excel(file, quiz_id: int, company: dict, member: dict, current_user: int):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=ERROR_NOT_EXCEL_FORMAT)
    file_data = await file.read()
    wb = load_workbook((io.BytesIO(file_data)), data_only=True)
    sheet = wb.active

    quiz_name = sheet['A1'].value
    quiz_title = sheet['A2'].value
    quiz_description = sheet['A3'].value
    quiz_frequency = sheet['A4'].value
    quiz_data = QuizUpdateModel(
        quiz_name=quiz_name,
        quiz_description=quiz_description,
        quiz_title=quiz_title,
        quiz_frequency=quiz_frequency
    )
    await quizzes_service.update_quiz(quiz_id, company, quiz_data, member, current_user)
    await question_service.remove_questions_for_quiz(quiz_id, company, member, current_user)
    await create_questions_from_excel(quiz_id, sheet, member, company, current_user)


async def create_questions_from_excel(quiz_id: int, sheet, member: dict, company: dict, current_user: int):
    quiz = await quizzes_service.get_quiz_by_id(quiz_id)
    for row in sheet.iter_rows(min_row=5, values_only=True):
        question_text, *question_answers, question_correct_answer = row

        question_data = QuestionCreateModel(
            question_text=question_text,
            question_answers=question_answers,
            question_correct_answer=question_correct_answer
        )

        await question_service.create_question(quiz, question_data, member, company, current_user)
